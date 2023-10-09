"""
erro provavelmente esta no read_excel
"""
from abstra.forms import *
import datetime
import pandas as pd
import openpyxl


start = (
    Page()
    .display("Let's start the Exam!")
    .read("First, what's your name?", key="name")
    .run()
)
name = start["name"]

n = datetime.datetime.now
n1 = n()

exam = pd.read_excel("src/files/Simulado-Exemplo.xlsx")
questions = exam["Question"].tolist()
optionA = exam["a)"].tolist()
optionB = exam["b)"].tolist()
optionC = exam["c)"].tolist()
optionD = exam["d)"].tolist()
optionE = exam["e)"].tolist()
answers = exam["answer"].tolist()

examPage = Page()

for index, question in enumerate(questions):
    examPage = examPage.read_cards(
        question,
        [
            {"title": "a)", "description": optionA[index]},
            {"title": "b)", "description": optionB[index]},
            {"title": "c)", "description": optionC[index]},
            {"title": "d)", "description": optionD[index]},
            {"title": "e)", "description": optionE[index]},
        ],
        key=f"answer{index+1}",
    )

studentAnswerSheet = examPage.run()
n2 = n()

examTime = n2 - n1
rightAnswers = 0
wrongAnswers = 0

wrkbk = openpyxl.Workbook()
sh = wrkbk.create_sheet(f"{name} AnswerSheet", 0)
sh.cell(row=1, column=1).value = "Name:"
sh.cell(row=1, column=2).value = f"{name}"
sh.cell(row=2, column=1).value = "Duration"
sh.cell(row=2, column=2).value = f"{examTime}"
sh.cell(row=3, column=1).value = "Score:"
sh.cell(row=5, column=1).value = "Question"
sh.cell(row=5, column=2).value = "Correct answer"
sh.cell(row=5, column=3).value = "Your answer"

for index, studentAnswer in enumerate(studentAnswerSheet.values()):
    if studentAnswer["title"] == answers[index]:
        rightAnswers = rightAnswers + 1
    else:
        wrongAnswers = wrongAnswers + 1
    sh.cell(row=index + 6, column=1).value = index + 1
    sh.cell(row=index + 6, column=2).value = answers[index]
    sh.cell(row=index + 6, column=3).value = studentAnswer["title"]

score = rightAnswers * 10 / (rightAnswers + wrongAnswers)
sh.cell(row=3, column=2).value = f"{score}"

fname = f"/tmp/{name}_answer_sheet.xlsx"
wrkbk.save(fname)

Page().display(f"Your Score: {score}").display(
    f"You made the exam in: {examTime}"
).display_file(fname, download_text="Check out here your exam results!").run()
